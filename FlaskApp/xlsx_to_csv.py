from openpyxl import Workbook, load_workbook
# from excel2csv import csv_from_excel
import openpyxl.utils.dataframe
import csv
import os
import pandas as pd
from itertools import islice
def csv_from_excel(file_name, DIR):
    s_wb = load_workbook(DIR+file_name)
    s_sh = s_wb.get_sheet_by_name('Data Sheet')
    d_wb = Workbook()
    d_sh = d_wb.get_sheet_by_name("Sheet")

    s_r1 = s_sh['A16':'K17']
    d_r1 = d_sh['A1':'K2']
    for s_r, d_r in zip(s_r1, d_r1):
        for s_c, d_c in zip(s_r, d_r):
            d_c.value = s_c.value
    s_r2 = s_sh['A26':'K31']
    d_r2 = d_sh['A3':'K8']
    for s_r, d_r in zip(s_r2, d_r2):
        for s_c, d_c in zip(s_r, d_r):
            d_c.value = s_c.value
    s_r3 = s_sh['A57':'K60']
    d_r3 = d_sh['A9':'K12']
    for s_r, d_r in zip(s_r3, d_r3):
        for s_c, d_c in zip(s_r, d_r):
            d_c.value = s_c.value

    s_r4 = s_sh['A62':'K65']
    d_r4 = d_sh['A13':'K16']
    for s_r, d_r in zip(s_r4, d_r4):
        for s_c, d_c in zip(s_r, d_r):
            d_c.value = s_c.value
    s_r5 = s_sh['A67':'K70']
    d_r5 = d_sh['A17':'K20']
    for s_r, d_r in zip(s_r5, d_r5):
        for s_c, d_c in zip(s_r, d_r):
            d_c.value = s_c.value

    data = d_sh.values
    cols = next(data)[1:]
    data = list(data)
    idx = [r[0] for r in data]
    data = (islice(r, 1, None) for r in data)
    df = pd.DataFrame(data, index=idx, columns=cols)
    df = df.transpose()
    df.to_csv(DIR+file_name.split(".")[0]+'.csv')
    # print(df.head())
    # d_sh = openpyxl.utils.dataframe.dataframe_to_rows(df, index=True, header=True)
    # # df.to_excel(DIR+'csv/'+file_name)
    #
    # temp_name = DIR+file_name.split(".")[0]+".csv"
    # # csv_from_excel(temp_name, ['Sheet'])
    # your_csv_file = open(temp_name, 'w')
    # wr = csv.writer(your_csv_file, quoting=csv.QUOTE_ALL)
    #
    # for row in d_sh.iter_rows():
    #     wr.writerow(cell.value for cell in row)
    # your_csv_file.close()
    # d_wb.save(DIR+"Edited/"+file_name)


DIR = 'C:/Users/Rajesh Rao/version_control/Stock_Screener_Data/Fundamental_Data/'
os.chdir(DIR)
files = filter(os.path.isfile, os.listdir(DIR))
files = [os.path.join(DIR, f) for f in files] # add path to each file
files.sort(key=lambda x: os.path.getmtime(x))

for cur_file in files:
    split_name = cur_file.split("/")
    name = split_name[-1]
    print(name)
    csv_from_excel(name, DIR)
    os.remove(DIR+name)
